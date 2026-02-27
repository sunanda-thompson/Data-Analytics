# ============================================================
#  MAGENTO DATA ANALYTICS — COMPLETE PYTHON REFERENCE
#  Resource Innovations | Data Analytics Coordinator Portfolio
# ============================================================
#  REQUIREMENTS: pip install pandas openpyxl
#  PYTHON:       3.8+
#
#  PURPOSE: Full pipeline from synthetic data generation →
#           database creation → validation → transformation →
#           settlement-ready output files
#
#  HOW TO USE:
#    Run each SECTION in order, or run the whole file at once.
#    Every transformation is explained with WHY, not just HOW.
#    In a real role, Sections 1-2 are replaced by actual
#    Magento API calls or CSV exports.
#
#  SECTIONS:
#    1. Setup & imports
#    2. Generate synthetic raw data (simulates Magento export)
#    3. Load data into SQLite (in-memory database)
#    4. Validation queries — find all data quality problems
#    5. Cleaning transformations — fix every problem found
#    6. Reconciliation — Magento vs payment processor
#    7. Build settlement-ready dataset
#    8. Export to CSV, JSON, and XLSX
#    9. Summary report
# ============================================================


# ============================================================
# SECTION 1: SETUP & IMPORTS
# ============================================================

import sqlite3          # built-in SQL database — no install needed
import pandas as pd     # data manipulation: pip install pandas
import json             # built-in: convert data to JSON format
import random           # built-in: generate random test data
import datetime         # built-in: work with dates and times
import os               # built-in: file paths and directories
from openpyxl import Workbook                         # Excel file creation
from openpyxl.styles import Font, PatternFill, Alignment   # Excel formatting
from openpyxl.utils import get_column_letter          # column letter helper

# Set a random seed so results are the same every run
random.seed(42)

# Create folder structure for outputs
for folder in ["data/raw", "data/processed", "data/output"]:
    os.makedirs(folder, exist_ok=True)

print("=" * 65)
print("  MAGENTO DATA ANALYTICS PIPELINE")
print("  Resource Innovations — Data Analytics Coordinator Project")
print("=" * 65)


# ============================================================
# SECTION 2: GENERATE SYNTHETIC RAW DATA
# ============================================================
# In the real role, this data would come from:
#   - Magento REST API:  GET /rest/V1/orders
#   - Commerce Manager: Admin → Sales → Orders → Export
#   - Authorize.net:    Reports → Transaction Detail → Export
#
# We generate it here with INTENTIONAL DATA QUALITY ISSUES
# to simulate what you actually receive from these systems.
# Issues injected:
#   1. Duplicate order IDs (export job ran twice)
#   2. SKU format inconsistencies (entered by different people)
#   3. Mixed tax formats (config change mid-year)
#   4. Currency stored as text with $ sign
#   5. Missing invoice numbers (~10% of orders)
#   6. Orphan transactions (payments with no Magento order)
# ============================================================

print("\n[SECTION 2] Generating synthetic raw data...")

# ── Customer data ────────────────────────────────────────────
# 50 customer records — the kind of data in Commerce Manager's
# customer table. Each has an ID, name, email, state, signup
# date, and loyalty tier.

first_names = ["Jordan","Taylor","Morgan","Casey","Riley","Drew","Avery","Quinn",
               "Skyler","Alex","Blake","Cameron","Dana","Emery","Frankie","Glen",
               "Hayden","Jamie","Kendall","Lee"]
last_names  = ["Smith","Johnson","Williams","Brown","Jones","Garcia","Miller",
               "Davis","Martinez","Wilson","Anderson","Thomas","Jackson","White",
               "Harris","Martin","Thompson","Moore","Young","Allen"]
states      = ["CA","TX","NY","FL","WA","OR","CO","IL","GA","AZ"]

customers = []
for i in range(1, 51):
    customers.append({
        "customer_id":  f"CUST-{i:04d}",
        "first_name":   random.choice(first_names),
        "last_name":    random.choice(last_names),
        "email":        f"user{i}@example.com",
        "state":        random.choice(states),
        # Date stored as ISO 8601 from the start — customers table is clean
        "created_at":   (datetime.date(2024, 1, 1) +
                         datetime.timedelta(days=random.randint(0, 180))).isoformat(),
        "loyalty_tier": random.choice(["bronze", "silver", "gold", "None"])
    })

df_customers = pd.DataFrame(customers)

# ── Orders data (intentionally messy) ───────────────────────
# SKU pool: same product appears under different formats.
# This mirrors reality — different people typed them differently,
# or a data migration didn't standardize the format.
sku_pool = [
    "SKU-LED-001",   "sku_led_001",    # ISSUE: same product, two formats
    "SKU-THERM-002", "THERM002",        # ISSUE: missing 'SKU-' prefix
    "SKU-SMART-003",                    # clean
    "SKU-AUDIT-004",                    # clean
    "SKU-HVAC-005",  "hvac-005",        # ISSUE: lowercase variant
    "SKU-REBATE-006",                   # clean
]
statuses = ["complete", "processing", "pending", "closed", "canceled"]

orders = []
base_date = datetime.datetime(2024, 6, 1)

for i in range(1, 101):
    cust       = random.choice(customers)
    subtotal   = round(random.uniform(20, 500), 2)

    # ISSUE: Two different tax recording approaches in same export.
    # About half the orders were created when Magento used combined tax.
    # The other half were created after a config change to itemized tax.
    use_combined = random.random() < 0.5
    state_tax    = round(subtotal * 0.06, 2)    # 6% state tax
    county_tax   = round(subtotal * 0.015, 2)   # 1.5% county tax
    combined     = round(state_tax + county_tax, 2) if use_combined else None

    shipping     = round(random.uniform(0, 25), 2)
    discount     = round(random.uniform(0, subtotal * 0.1), 2)
    grand_total  = round(
        subtotal + (combined if combined else state_tax + county_tax)
        + shipping - discount, 2
    )

    orders.append({
        "order_id":       f"ORD-{i:05d}",
        "customer_id":    cust["customer_id"],
        # ISSUE: Date stored in US format, not ISO 8601
        # Magento defaults to MM/DD/YYYY — needs conversion
        "order_date":     (base_date + datetime.timedelta(
                              days=random.randint(0, 90),
                              hours=random.randint(0, 23),
                              minutes=random.randint(0, 59)
                          )).strftime("%m/%d/%Y %H:%M"),
        "sku":            random.choice(sku_pool),
        "qty":            random.randint(1, 5),
        # ISSUE: Subtotal exported as a formatted currency string
        # e.g. "$215.08" instead of the number 215.08
        "subtotal":       f"${subtotal:,.2f}",
        # Tax fields: one format or the other, never both
        "state_tax":      state_tax if not use_combined else None,
        "county_tax":     county_tax if not use_combined else None,
        "combined_tax":   combined,
        "shipping":       shipping,
        "discount":       discount,
        "grand_total":    grand_total,
        "status":         random.choice(statuses),
        "payment_method": random.choice(["authorizenet", "paypal", "free"]),
        # ISSUE: ~10% of orders have no invoice_number
        # This happens when the billing step didn't complete
        "invoice_number": f"INV-{i:05d}" if random.random() > 0.1 else None,
    })

# ISSUE: Inject 5 duplicate rows (simulates a re-export overlap)
# In practice: someone exported June 1-30, then exported June 15 - July 15,
# causing orders from June 15-30 to appear in both exports.
dup_indices = random.sample(range(100), 5)
for idx in dup_indices:
    orders.append(orders[idx].copy())

df_orders = pd.DataFrame(orders)

# ── Payment transaction data (Authorize.net style) ───────────
# Only complete/processing orders would have a payment transaction.
# We generate transactions only for those, then add orphans.
transactions = []
settled_orders = (
    df_orders[df_orders["status"].isin(["complete", "processing"])]
    .drop_duplicates("order_id")
)

for _, row in settled_orders.iterrows():
    # Authorize.net charges 2.9% of the transaction + $0.30 flat fee
    fee = round(row["grand_total"] * 0.029 + 0.30, 2)
    net = round(row["grand_total"] - fee, 2)

    transactions.append({
        "transaction_id":  f"TXN-{random.randint(100000, 999999)}",
        "order_id":        row["order_id"],
        # Settlements typically arrive 2 business days after the order
        "settle_date":     (
            datetime.datetime.strptime(row["order_date"], "%m/%d/%Y %H:%M")
            + datetime.timedelta(days=2)
        ).strftime("%Y-%m-%d"),
        "gross_amount":    row["grand_total"],
        "processor_fee":   fee,
        "net_amount":      net,
        # Not every transaction settles: some get voided or refunded
        "status":          random.choice(["settled", "settled", "settled",
                                          "voided", "refunded"]),
        "auth_code":       f"AUTH{random.randint(10000, 99999)}",
    })

# ISSUE: Inject 3 orphan transactions — money the processor has
# on record but no matching Magento order exists.
# This could mean: deleted orders, orders from another system,
# or (worst case) fraudulent charges.
for i in range(3):
    transactions.append({
        "transaction_id": f"TXN-ORPHAN-{i}",
        "order_id":       f"ORD-GHOST-{i}",   # these IDs don't exist in Magento
        "settle_date":    "2024-09-01",
        "gross_amount":   round(random.uniform(50, 200), 2),
        "processor_fee":  5.00,
        "net_amount":     round(random.uniform(45, 195), 2),
        "status":         "settled",
        "auth_code":      f"AUTH99{i}",
    })

df_transactions = pd.DataFrame(transactions)

# ── Save raw CSVs ────────────────────────────────────────────
# These files represent the raw exports before ANY cleaning.
# Always save the originals — never overwrite them.
df_customers.to_csv("data/raw/magento_customers.csv", index=False)
df_orders.to_csv("data/raw/magento_orders.csv", index=False)
df_transactions.to_csv("data/raw/payment_transactions.csv", index=False)

print(f"  ✅ {len(df_customers)} customers saved to data/raw/magento_customers.csv")
print(f"  ✅ {len(df_orders)} orders saved  (includes {len(df_orders)-100} duplicate rows)")
print(f"  ✅ {len(df_transactions)} transactions saved (includes 3 orphan records)")


# ============================================================
# SECTION 3: LOAD DATA INTO SQLite
# ============================================================
# SQLite is a lightweight database that runs entirely in memory
# (":memory:") or as a single file. It's perfect for:
#   - Running SQL validation queries on CSV exports
#   - Staging data before transformation
#   - Prototyping queries before moving to MySQL/PostgreSQL
#
# In production: you'd connect to the actual Magento MySQL DB
# or a data warehouse instead of SQLite.
# ============================================================

print("\n[SECTION 3] Loading raw data into SQLite...")

# Create an in-memory SQLite database
# ":memory:" means the database lives only in RAM — no file created
conn = sqlite3.connect(":memory:")
conn.execute("PRAGMA journal_mode=WAL")  # improves write performance

# Load each dataframe as a SQL table
# if_exists="replace" drops and recreates the table each run
df_orders.to_sql("magento_orders",       conn, index=False, if_exists="replace")
df_customers.to_sql("magento_customers", conn, index=False, if_exists="replace")
df_transactions.to_sql("payment_transactions", conn, index=False, if_exists="replace")

def run_sql(sql, label=""):
    """
    Helper: execute a SQL query and return results as a DataFrame.
    Prints the label and results to the console for review.
    """
    df = pd.read_sql(sql, conn)
    if label:
        print(f"\n  ── {label}")
        if df.empty:
            print("     ✅ No issues found.")
        else:
            # Indent the output for readability
            for line in df.to_string(index=False).split("\n"):
                print(f"     {line}")
    return df

print("  ✅ Tables loaded:")
counts = run_sql("""
    SELECT 'magento_orders' AS tbl, COUNT(*) AS rows FROM magento_orders
    UNION ALL SELECT 'magento_customers', COUNT(*) FROM magento_customers
    UNION ALL SELECT 'payment_transactions', COUNT(*) FROM payment_transactions
""")


# ============================================================
# SECTION 4: VALIDATION — FIND ALL DATA QUALITY PROBLEMS
# ============================================================
# These queries run BEFORE any cleaning.
# Document every issue found so there is an audit trail.
# In the real role, you'd share these findings with the
# finance team before proceeding to transformation.
# ============================================================

print("\n" + "=" * 65)
print("  SECTION 4: VALIDATION QUERIES")
print("=" * 65)

issues_log = []  # We'll collect all issues here for the output files


# ── 4A. Duplicate detection ──────────────────────────────────
# GROUP BY order_id, count how many times each appears.
# HAVING COUNT(*) > 1 filters to only the duplicates.
print("\n[4A] Checking for duplicate order IDs...")
dupes = run_sql("""
    SELECT
        order_id,
        COUNT(*) AS occurrences
    FROM magento_orders
    GROUP BY order_id
    HAVING COUNT(*) > 1
    ORDER BY occurrences DESC
""", "Duplicate order IDs found:")

# Log each duplicate to the issues list
for _, row in dupes.iterrows():
    issues_log.append({
        "issue_type":     "DUPLICATE_ORDER",
        "order_id":       row["order_id"],
        "transaction_id": None,
        "detail":         f"order_id appears {row['occurrences']}x in export",
        "action_required": "Keep first occurrence — delete remaining rows"
    })
print(f"  → {len(dupes)} order IDs have duplicates")


# ── 4B. Missing field audit ──────────────────────────────────
# CASE WHEN ... IS NULL THEN 1 ELSE 0 END counts NULLs per column.
# SUM() across all rows gives the total null count per column.
print("\n[4B] Auditing for missing (NULL) fields...")
run_sql("""
    SELECT
        SUM(CASE WHEN order_id       IS NULL THEN 1 ELSE 0 END) AS null_order_id,
        SUM(CASE WHEN customer_id    IS NULL THEN 1 ELSE 0 END) AS null_customer_id,
        SUM(CASE WHEN sku             IS NULL THEN 1 ELSE 0 END) AS null_sku,
        SUM(CASE WHEN grand_total     IS NULL THEN 1 ELSE 0 END) AS null_grand_total,
        SUM(CASE WHEN invoice_number  IS NULL THEN 1 ELSE 0 END) AS null_invoice_number,
        SUM(CASE WHEN payment_method  IS NULL THEN 1 ELSE 0 END) AS null_payment_method,
        COUNT(*) AS total_rows
    FROM magento_orders
""", "NULL count per column:")

missing_inv = run_sql("""
    SELECT order_id, customer_id, status, grand_total, invoice_number
    FROM magento_orders
    WHERE invoice_number IS NULL
    ORDER BY order_id
""", "Orders missing invoice_number:")

for _, row in missing_inv.iterrows():
    issues_log.append({
        "issue_type":     "MISSING_INVOICE",
        "order_id":       row["order_id"],
        "transaction_id": None,
        "detail":         "invoice_number is NULL — cannot submit for payment",
        "action_required": "Request invoice number from finance/billing team"
    })
print(f"  → {len(missing_inv)} orders are missing an invoice number")


# ── 4C. Tax format inconsistency ────────────────────────────
# Count how many orders use each tax format.
# We'll need to unify these in Section 5.
print("\n[4C] Checking tax field format consistency...")
run_sql("""
    SELECT
        SUM(CASE WHEN combined_tax IS NOT NULL                         THEN 1 ELSE 0 END) AS combined_format,
        SUM(CASE WHEN state_tax IS NOT NULL AND county_tax IS NOT NULL THEN 1 ELSE 0 END) AS itemized_format,
        COUNT(*) AS total_orders
    FROM magento_orders
""", "Tax format distribution:")

# Show sample of each format side by side:
run_sql("""
    SELECT order_id, subtotal, combined_tax, state_tax, county_tax, grand_total
    FROM magento_orders
    WHERE combined_tax IS NOT NULL
    LIMIT 3
""", "Sample: combined tax format (3 rows):")

run_sql("""
    SELECT order_id, subtotal, combined_tax, state_tax, county_tax, grand_total
    FROM magento_orders
    WHERE state_tax IS NOT NULL
    LIMIT 3
""", "Sample: itemized tax format (3 rows):")


# ── 4D. SKU format inconsistency ────────────────────────────
# Find SKU values that map to the same normalized code.
# These are the same product but entered differently.
print("\n[4D] Checking for inconsistent SKU formats...")
run_sql("""
    SELECT
        UPPER(REPLACE(REPLACE(sku, '_', '-'), ' ', '-'))       AS normalized_sku,
        COUNT(DISTINCT sku)                                     AS raw_variant_count,
        GROUP_CONCAT(DISTINCT sku)                              AS raw_variants,
        COUNT(*)                                                AS total_orders
    FROM magento_orders
    GROUP BY UPPER(REPLACE(REPLACE(sku, '_', '-'), ' ', '-'))
    HAVING COUNT(DISTINCT sku) > 1
""", "SKU variants that should be the same product:")


# ── 4E. Currency format check ────────────────────────────────
# The subtotal column contains strings like "$215.08" not numbers.
# Verify all can be converted cleanly.
print("\n[4E] Checking currency format (subtotal column)...")
sample = run_sql("""
    SELECT
        order_id,
        subtotal AS raw_subtotal,
        CAST(REPLACE(REPLACE(subtotal,'$',''),',','') AS REAL) AS numeric_subtotal
    FROM magento_orders
    LIMIT 5
""", "Sample: subtotal as raw string vs converted number:")


# ── 4F. Date format check ────────────────────────────────────
# Dates in 'MM/DD/YYYY HH:MM' format need to become ISO 8601
print("\n[4F] Checking date format...")
run_sql("""
    SELECT
        order_id,
        order_date AS raw_date,
        -- Reconstruct to ISO format using SUBSTR:
        -- Original: '06/22/2024 12:00'
        -- SUBSTR(order_date,7,4) → '2024'   (year, starting at char 7, 4 chars long)
        -- SUBSTR(order_date,1,2) → '06'     (month)
        -- SUBSTR(order_date,4,2) → '22'     (day)
        SUBSTR(order_date,7,4)||'-'||SUBSTR(order_date,1,2)||'-'||SUBSTR(order_date,4,2) AS iso_date
    FROM magento_orders
    LIMIT 5
""", "Sample: raw date vs ISO 8601 conversion:")


# ── 4G. Orphan transaction check ────────────────────────────
# Look for transactions in the processor file that have no
# matching order in Magento.
print("\n[4G] Checking for orphan transactions...")
orphans = run_sql("""
    SELECT
        t.transaction_id,
        t.order_id,
        t.gross_amount,
        t.status,
        t.settle_date
    FROM payment_transactions t
    LEFT JOIN magento_orders o ON t.order_id = o.order_id
    WHERE o.order_id IS NULL    -- LEFT JOIN found no match → orphan
      AND t.status = 'settled'
""", "Orphan transactions (settled but no Magento order):")

for _, row in orphans.iterrows():
    issues_log.append({
        "issue_type":     "ORPHAN_TRANSACTION",
        "order_id":       row["order_id"],
        "transaction_id": row["transaction_id"],
        "detail":         f"${row['gross_amount']} settled — no matching Magento order",
        "action_required": "Escalate to finance team for investigation"
    })
print(f"  → {len(orphans)} orphan transactions found")

# Save the issues log for later
df_issues = pd.DataFrame(issues_log)
print(f"\n  Total issues logged: {len(df_issues)}")


# ============================================================
# SECTION 5: CLEANING — TRANSFORM THE RAW DATA
# ============================================================
# Now we fix every problem found in Section 4.
# We NEVER modify the raw tables — we create a new clean table.
# Each transformation is done in Python (pandas) and mirrored
# by the equivalent SQL in the .sql file.
# ============================================================

print("\n" + "=" * 65)
print("  SECTION 5: DATA CLEANING & TRANSFORMATION")
print("=" * 65)

# Reload the raw CSV to work from a fresh copy
orders = pd.read_csv("data/raw/magento_orders.csv")
print(f"\n  Starting with {len(orders)} raw order rows")


# ── Transform 1: Remove duplicates ──────────────────────────
# keep="first" retains the first occurrence of each order_id
# and marks later occurrences as duplicates to drop.
print("\n[T1] Deduplication...")
before = len(orders)
orders = orders.drop_duplicates(subset="order_id", keep="first")
after  = len(orders)
print(f"  Removed {before - after} duplicate rows → {after} unique orders remain")


# ── Transform 2: SKU normalization ──────────────────────────
# Rule: uppercase all characters, replace underscores with hyphens.
# This makes 'sku_led_001', 'SKU-LED-001', 'SKU_LED_001' all
# become the same canonical 'SKU-LED-001'.
print("\n[T2] SKU normalization...")

sku_before = orders["sku"].value_counts().to_dict()

orders["normalized_sku"] = (
    orders["sku"]
    .str.upper()           # 'sku_led_001' → 'SKU_LED_001'
    .str.strip()           # remove leading/trailing spaces
    .str.replace("_", "-", regex=False)  # 'SKU_LED_001' → 'SKU-LED-001'
    .str.replace(" ", "-", regex=False)  # handle any spaces in SKU codes
)

# Report: show all cases where normalization changed the value
changed = orders[orders["sku"] != orders["normalized_sku"]][["order_id", "sku", "normalized_sku"]]
print(f"  {len(changed)} orders had their SKU normalized:")
print(changed.to_string(index=False))


# ── Transform 3: Currency normalization ─────────────────────
# The subtotal column contains strings like "$215.08" or "$1,247.50".
# We need to:
#   1. Remove the $ sign
#   2. Remove any commas (for values over $1,000)
#   3. Convert to a float (decimal number) so math works
print("\n[T3] Currency normalization (subtotal)...")
sample_before = orders["subtotal"].head(3).tolist()

orders["subtotal_clean"] = (
    orders["subtotal"]
    .str.replace("$", "", regex=False)    # remove dollar sign
    .str.replace(",", "", regex=False)    # remove thousands comma
    .astype(float)                         # convert string to number
    .round(2)                              # enforce 2 decimal places
)

print(f"  Sample: {sample_before} → {orders['subtotal_clean'].head(3).tolist()}")

# Also round all other monetary fields for consistency
for col in ["shipping", "discount", "grand_total"]:
    orders[col] = orders[col].round(2)


# ── Transform 4: Tax field unification ──────────────────────
# Business rule: the client's system needs ONE total_tax field.
# Strategy:
#   - If combined_tax exists → use it directly
#   - If not → add state_tax + county_tax together
#   - Track which method was used in a 'tax_source' column
print("\n[T4] Tax field unification...")

def resolve_tax(row):
    """
    Resolve the tax amount from whichever format is present.
    Returns (total_tax, source_format).
    """
    if pd.notna(row["combined_tax"]):
        # Order was created when Magento used combined tax
        return round(row["combined_tax"], 2), "combined"
    elif pd.notna(row["state_tax"]) and pd.notna(row["county_tax"]):
        # Order was created when Magento used itemized tax
        return round(row["state_tax"] + row["county_tax"], 2), "itemized"
    else:
        # Neither field is populated — flag for review
        return None, "missing"

# apply() runs the function on every row, result_type="expand"
# splits the tuple return into two columns
tax_result = orders.apply(resolve_tax, axis=1, result_type="expand")
orders["total_tax"]  = tax_result[0]   # the unified tax amount
orders["tax_source"] = tax_result[1]   # 'combined', 'itemized', or 'missing'

tax_dist = orders["tax_source"].value_counts()
print(f"  combined: {tax_dist.get('combined',0)} orders")
print(f"  itemized: {tax_dist.get('itemized',0)} orders")
print(f"  missing:  {tax_dist.get('missing',0)} orders")


# ── Transform 5: Date normalization ─────────────────────────
# Convert 'MM/DD/YYYY HH:MM' to ISO 8601: 'YYYY-MM-DDTHH:MM:SSZ'
# pd.to_datetime() is smart enough to parse most date formats.
# errors="coerce" turns any unparseable dates into NaT (null) instead of crashing.
print("\n[T5] Date normalization...")
sample_before = orders["order_date"].head(3).tolist()

orders["order_date_iso"] = (
    pd.to_datetime(orders["order_date"], errors="coerce")
    .dt.strftime("%Y-%m-%dT%H:%M:%SZ")
)

bad_dates = orders["order_date_iso"].isna().sum()
print(f"  Converted {len(orders) - bad_dates} dates to ISO 8601")
print(f"  Unparseable dates: {bad_dates}")
print(f"  Before: {sample_before[0]}  →  After: {orders['order_date_iso'].iloc[0]}")


# ── Transform 6: Status code mapping ────────────────────────
# Magento uses plain English status labels.
# Microsoft Dynamics expects numeric codes.
# We also add a boolean flag: is this order eligible for payment submission?
print("\n[T6] Status code mapping...")

STATUS_MAP = {
    # magento_status: (dynamics_code, label, payment_eligible)
    "complete":   ("110", "INVOICED",    True),
    "processing": ("100", "IN_PROGRESS", True),
    "closed":     ("120", "CLOSED",      False),
    "pending":    ("50",  "PENDING",     False),
    "canceled":   ("999", "VOID",        False),
}

orders["dynamics_code"]      = orders["status"].map(lambda s: STATUS_MAP.get(s, ("000","UNKNOWN",False))[0])
orders["status_label"]        = orders["status"].map(lambda s: STATUS_MAP.get(s, ("000","UNKNOWN",False))[1])
orders["payment_eligible"]    = orders["status"].map(lambda s: STATUS_MAP.get(s, ("000","UNKNOWN",False))[2])

print(pd.crosstab(orders["status"], orders["dynamics_code"]).to_string())


# ── Transform 7: SKU → Incentive program mapping ─────────────
# Each product SKU corresponds to an energy efficiency program.
# The incentive_rate determines what percentage of the subtotal
# the utility company reimburses to Resource Innovations.
print("\n[T7] SKU → Incentive program mapping...")

INCENTIVE_MAP = {
    # sku: (program_name, reimbursement_rate)
    "SKU-LED-001":    ("ENERGY_EFF_LIGHTING", 0.15),   # 15% of subtotal
    "SKU-THERM-002":  ("SMART_THERMOSTAT",    0.20),   # 20%
    "SKU-SMART-003":  ("SMART_HOME_PROG",     0.18),   # 18%
    "SKU-AUDIT-004":  ("HOME_ENERGY_AUDIT",   0.25),   # 25%
    "SKU-HVAC-005":   ("HVAC_UPGRADE",        0.22),   # 22%
    "HVAC-005":       ("HVAC_UPGRADE",        0.22),   # same program, pre-normalized
    "SKU-REBATE-006": ("DIRECT_REBATE",       0.30),   # 30%
    "THERM002":       ("SMART_THERMOSTAT",    0.20),   # same program, pre-normalized
}

orders["incentive_program"] = orders["normalized_sku"].map(
    lambda s: INCENTIVE_MAP.get(s, ("UNMAPPED", 0))[0]
)
orders["incentive_rate"] = orders["normalized_sku"].map(
    lambda s: INCENTIVE_MAP.get(s, ("UNMAPPED", 0))[1]
)
# incentive_amount = subtotal × rate, rounded to 2 decimal places
orders["incentive_amount"] = (orders["subtotal_clean"] * orders["incentive_rate"]).round(2)

unmapped = orders[orders["incentive_program"] == "UNMAPPED"]
print(f"  Unmapped SKUs: {len(unmapped)}")
print(f"  Distribution:")
for prog, cnt in orders["incentive_program"].value_counts().items():
    print(f"    {prog}: {cnt} orders")


# ── Transform 8: Grand total cross-check ────────────────────
# Recompute grand_total from its component parts.
# If the recomputed value differs from the recorded value by more
# than $0.02, something is wrong with the order data.
print("\n[T8] Grand total recomputation check...")

orders["grand_total_check"] = (
    orders["subtotal_clean"] +
    orders["total_tax"].fillna(0) +      # fillna(0) treats NULL tax as $0
    orders["shipping"].fillna(0) -
    orders["discount"].fillna(0)
).round(2)

mismatches = orders[
    (orders["grand_total_check"] - orders["grand_total"]).abs() > 0.02
]
print(f"  Mismatches found: {len(mismatches)}")
if len(mismatches) == 0:
    print("  ✅ All grand totals reconcile correctly")
else:
    print(mismatches[["order_id","grand_total","grand_total_check"]].to_string(index=False))

# Clean up the helper column
orders.drop(columns=["grand_total_check"], inplace=True)

# Save the clean orders to a processed CSV
orders.to_csv("data/processed/orders_clean.csv", index=False)
print(f"\n  ✅ Saved cleaned orders: data/processed/orders_clean.csv ({len(orders)} rows)")


# ============================================================
# SECTION 6: RECONCILIATION — MAGENTO vs PAYMENT PROCESSOR
# ============================================================
# Load the clean orders and transactions into SQLite for
# the reconciliation queries. This mirrors what you'd do when
# comparing against the Authorize.net settlement file.
# ============================================================

print("\n" + "=" * 65)
print("  SECTION 6: RECONCILIATION CHECKS")
print("=" * 65)

# Load clean orders and transactions into SQLite
orders_clean = pd.read_csv("data/processed/orders_clean.csv")
transactions  = pd.read_csv("data/raw/payment_transactions.csv")
customers_df  = pd.read_csv("data/raw/magento_customers.csv")

orders_clean.to_sql("orders_clean",        conn, index=False, if_exists="replace")
transactions.to_sql("payment_transactions", conn, index=False, if_exists="replace")
customers_df.to_sql("magento_customers",    conn, index=False, if_exists="replace")


# ── 6A. High-level reconciliation ───────────────────────────
print("\n[6A] Magento vs Authorize.net — order count and revenue comparison:")
run_sql("""
    SELECT 'Magento (complete+processing)' AS source,
           COUNT(DISTINCT order_id)        AS order_count,
           ROUND(SUM(grand_total), 2)      AS total_revenue
    FROM orders_clean
    WHERE status IN ('complete','processing')

    UNION ALL

    SELECT 'Authorize.net (settled)',
           COUNT(DISTINCT order_id),
           ROUND(SUM(gross_amount), 2)
    FROM payment_transactions
    WHERE status = 'settled'
""", "Reconciliation summary:")


# ── 6B. Orphan transactions ──────────────────────────────────
print("\n[6B] Orphan transactions (in processor, not in Magento):")
run_sql("""
    SELECT t.transaction_id, t.order_id, t.gross_amount, t.status, t.settle_date
    FROM payment_transactions t
    LEFT JOIN orders_clean o ON t.order_id = o.order_id
    WHERE o.order_id IS NULL AND t.status = 'settled'
""", "Orphan transactions:")


# ── 6C. Unsettled Magento orders ─────────────────────────────
print("\n[6C] Magento orders with no matching settlement:")
run_sql("""
    SELECT o.order_id, o.status, o.grand_total, o.payment_method
    FROM orders_clean o
    LEFT JOIN payment_transactions t ON o.order_id = t.order_id
    WHERE o.payment_eligible = 1     -- SQLite stores True as 1
      AND t.order_id IS NULL
""", "Unsettled orders:")


# ── 6D. Amount mismatches ────────────────────────────────────
print("\n[6D] Amount mismatches (same order, different dollar amounts):")
run_sql("""
    SELECT
        o.order_id,
        o.grand_total    AS magento_total,
        t.gross_amount   AS processor_total,
        ROUND(ABS(o.grand_total - t.gross_amount), 2) AS discrepancy
    FROM orders_clean o
    INNER JOIN payment_transactions t ON o.order_id = t.order_id
    WHERE ABS(o.grand_total - t.gross_amount) > 0.01
    ORDER BY discrepancy DESC
""", "Amount mismatches:")


# ============================================================
# SECTION 7: BUILD SETTLEMENT-READY DATASET
# ============================================================
# Join clean orders + settled transactions + customer info.
# Only include orders that are:
#   1. Payment-eligible (complete or processing status)
#   2. Have a matching settled transaction in the processor file
# ============================================================

print("\n" + "=" * 65)
print("  SECTION 7: BUILD SETTLEMENT FILE")
print("=" * 65)

# Filter: only settled transactions
settled = transactions[transactions["status"] == "settled"].copy()
settled = settled.rename(columns={
    "transaction_id": "txn_id",
    "settle_date":    "settlement_date",
    "processor_fee":  "processing_fee",
    "net_amount":     "net_settled_amount",
})

# Filter: only payment-eligible orders
eligible = orders_clean[orders_clean["payment_eligible"] == True].copy()

# Join 1: eligible orders + settled transactions (INNER JOIN = must match both)
# An inner join keeps only rows where the order_id exists in BOTH tables.
settlement = eligible.merge(
    settled[["order_id", "txn_id", "settlement_date", "processing_fee", "net_settled_amount"]],
    on="order_id",
    how="inner"   # only keep orders that have a settled transaction
)

# Join 2: add customer information (LEFT JOIN = keep all orders even if no customer found)
settlement = settlement.merge(
    customers_df[["customer_id", "first_name", "last_name", "email", "state"]],
    on="customer_id",
    how="left"
)

# Select and rename columns for the final output schema
settlement_final = settlement[[
    "order_id", "invoice_number", "order_date_iso", "settlement_date",
    "customer_id", "first_name", "last_name", "email", "state",
    "normalized_sku", "incentive_program", "qty",
    "subtotal_clean", "total_tax", "tax_source", "shipping", "discount",
    "grand_total", "incentive_rate", "incentive_amount",
    "processing_fee", "net_settled_amount",
    "dynamics_code", "status_label", "txn_id"
]].rename(columns={
    "order_date_iso":  "order_timestamp",
    "subtotal_clean":  "subtotal",
    "normalized_sku":  "sku_normalized",
    "dynamics_code":   "dynamics_status_code",
})

print(f"\n  Settlement records: {len(settlement_final)}")
print(f"  Total gross revenue:    ${settlement_final['grand_total'].sum():,.2f}")
print(f"  Total incentives due:   ${settlement_final['incentive_amount'].sum():,.2f}")
print(f"  Total processor fees:   ${settlement_final['processing_fee'].sum():,.2f}")
print(f"  Net to client:          ${settlement_final['net_settled_amount'].sum():,.2f}")


# ============================================================
# SECTION 8: EXPORT OUTPUT FILES
# ============================================================
# Generate three output formats — each serves a different
# downstream system or audience.
# ============================================================

print("\n" + "=" * 65)
print("  SECTION 8: EXPORTING OUTPUT FILES")
print("=" * 65)


# ── Output 1: CSV ────────────────────────────────────────────
# Best for: ETL pipelines, internal data exchange, bulk imports
# Format: plain text, one row per order, no formatting
csv_path = "data/output/settlement_ready.csv"
settlement_final.to_csv(csv_path, index=False)
print(f"\n  [CSV] Saved: {csv_path}")


# ── Output 2: JSON ───────────────────────────────────────────
# Best for: REST API uploads to Microsoft Dynamics or ARIBA
# Format: structured key-value pairs, machine-readable
json_path = "data/output/settlement_ready.json"

# Build a JSON payload with a summary header and the order records
json_payload = {
    "export_timestamp":   datetime.datetime.now().strftime("%Y-%m-%dT%H:%M:%SZ"),
    "source_system":      "Magento / Commerce Manager",
    "target_system":      "Microsoft Dynamics",
    "record_count":       len(settlement_final),
    "summary": {
        "total_gross_revenue":     round(float(settlement_final["grand_total"].sum()), 2),
        "total_incentive_payable": round(float(settlement_final["incentive_amount"].sum()), 2),
        "total_processing_fees":   round(float(settlement_final["processing_fee"].sum()), 2),
        "net_to_client":           round(float(settlement_final["net_settled_amount"].sum()), 2),
    },
    # orient="records" creates a list of {column: value} dicts — one per row
    "orders": json.loads(settlement_final.to_json(orient="records", date_format="iso"))
}

with open(json_path, "w") as f:
    json.dump(json_payload, f, indent=2, default=str)
print(f"  [JSON] Saved: {json_path}")


# ── Output 3: XLSX (formatted invoice report) ────────────────
# Best for: finance team, account managers, client presentations
# Format: Excel with 4 tabs, colored headers, auto-sized columns
xlsx_path = "data/output/invoice_report.xlsx"

with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:

    # ── Tab 1: Settlement Detail (all records) ────────────────
    settlement_final.to_excel(writer, sheet_name="Settlement Detail", index=False)

    # ── Tab 2: Program Summary (grouped by incentive program) ──
    program_summary = (
        settlement_final
        .groupby("incentive_program")
        .agg(
            order_count=("order_id", "count"),
            total_subtotal=("subtotal", "sum"),
            total_tax=("total_tax", "sum"),
            total_incentive=("incentive_amount", "sum"),
            net_revenue=("net_settled_amount", "sum")
        )
        .round(2)
        .reset_index()
        .sort_values("net_revenue", ascending=False)
    )
    program_summary.to_excel(writer, sheet_name="Program Summary", index=False)

    # ── Tab 3: Monthly Summary (grouped by settlement month) ───
    settlement_final["month"] = (
        pd.to_datetime(settlement_final["settlement_date"], errors="coerce")
        .dt.to_period("M")
        .astype(str)
    )
    monthly = (
        settlement_final
        .groupby("month")
        .agg(
            transactions=("order_id", "count"),
            gross_revenue=("grand_total", "sum"),
            processing_fees=("processing_fee", "sum"),
            incentive_payable=("incentive_amount", "sum"),
            net_to_client=("net_settled_amount", "sum")
        )
        .round(2)
        .reset_index()
    )
    monthly.to_excel(writer, sheet_name="Monthly Summary", index=False)

    # ── Tab 4: Issues Log (for finance team) ──────────────────
    df_issues.to_excel(writer, sheet_name="Issues Log", index=False)

    # ── Apply formatting to all tabs ──────────────────────────
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", name="Arial")
    center_align = Alignment(horizontal="center")

    for sheet_name in writer.sheets:
        ws = writer.sheets[sheet_name]
        # Style the header row
        for cell in ws[1]:
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = center_align
        # Auto-size all columns
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 3, 40)
        ws.freeze_panes = "A2"  # freeze header row when scrolling

print(f"  [XLSX] Saved: {xlsx_path} (4 tabs)")

# ── Output 4: Validation flags CSV ───────────────────────────
# Standalone issues file for quick handoff to the finance team
flags_path = "data/output/validation_flags.csv"
df_issues.to_csv(flags_path, index=False)
print(f"  [CSV] Saved: {flags_path} (issues log)")


# ============================================================
# SECTION 9: SUMMARY REPORT
# ============================================================

print("\n" + "=" * 65)
print("  PIPELINE COMPLETE — FINAL SUMMARY")
print("=" * 65)

print(f"""
  RAW DATA (from Magento + Authorize.net):
  ├── {len(df_orders)} order rows (including duplicates)
  ├── {len(df_customers)} customer records
  └── {len(df_transactions)} payment transactions

  ISSUES FOUND:
  ├── {len(df_issues[df_issues['issue_type']=='DUPLICATE_ORDER'])} duplicate order IDs
  ├── {len(df_issues[df_issues['issue_type']=='MISSING_INVOICE'])} orders missing invoice number
  └── {len(df_issues[df_issues['issue_type']=='ORPHAN_TRANSACTION'])} orphan transactions (${orphans['gross_amount'].sum():,.2f})

  TRANSFORMATIONS APPLIED:
  ├── Deduplication: {len(df_orders) - len(orders)} rows removed
  ├── SKU normalization: {len(changed)} SKUs standardized
  ├── Tax unification: combined + itemized → single total_tax
  ├── Currency: "$215.08" string → 215.08 float
  ├── Dates: MM/DD/YYYY → ISO 8601
  └── Status codes: complete → 110, processing → 100, etc.

  SETTLEMENT FILE:
  ├── {len(settlement_final)} payment-eligible settled records
  ├── Gross revenue:      ${settlement_final['grand_total'].sum():,.2f}
  ├── Incentives payable: ${settlement_final['incentive_amount'].sum():,.2f}
  ├── Processor fees:     ${settlement_final['processing_fee'].sum():,.2f}
  └── Net to client:      ${settlement_final['net_settled_amount'].sum():,.2f}

  OUTPUT FILES:
  ├── data/output/settlement_ready.csv   (upload-ready flat file)
  ├── data/output/settlement_ready.json  (API payload for Dynamics)
  ├── data/output/invoice_report.xlsx    (4-tab client report)
  └── data/output/validation_flags.csv   (exceptions for finance team)
""")

conn.close()
